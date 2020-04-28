import requests, datetime
from bs4 import BeautifulSoup as bs
from openpyxl import Workbook
from openpyxl.styles import Alignment
session_requests = requests.session()
loginUrl = 'https://springtour.app/Ajax/loginadmin'
payload = {
    'account': 'xxxxx',
    'password': 'xxxxx'
}
headers = {
    'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
    'Origin': 'https://springtour.app',
    'Referer': 'https://springtour.app/Main/loginAdmin',
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.142 Safari/537.36',
    'X-Requested-With': 'XMLHttpRequest'
}
date = str(datetime.date.today()).replace('-', '')
r = session_requests.post(loginUrl, data=payload, headers=headers)
if r.status_code == requests.codes.ok:
    print('Login Success!')
    hotelCount = 240
    # for i in range(28, 33):
    #     targetUrl = 'https://springtour.app/Gov/statis/{}'.format(i)
    #     r = session_requests.get(targetUrl, headers = dict(referer = targetUrl))
    #     soup = bs(r.text, 'html.parser')
    #     hotelInfos = soup.find_all('td', {'class': 'cal6'})
    #     for hotelInfo in hotelInfos:
    #         hotelLink = hotelInfo.find('a').get('href')
    orderManual = '5566'
    hotelUrl = 'https://springtour.app/Gov/statisdetail/{}'.format(orderManual)
    r = session_requests.get(hotelUrl)
    soup = bs(r.text, 'html.parser')
    hotelNameTag = soup.select('#user_panel p')[0]
    hotelName = hotelNameTag.contents[0]
    print(hotelName)
    orderNumbers = soup.find_all('td', {'class': 'cal1'})
    orderInfos = soup.find_all('td', {'class': 'cal6'})
    orderNumbersLst = []
    orderInfosLst = []
    wb = Workbook()
    for orderNumber in orderNumbers:
        orderNumber = orderNumber.text
        orderNumbersLst.append(orderNumber)
    for orderInfo in orderInfos:
        orderLink = orderInfo.find('a').get('href')
        orderUrl = 'https://springtour.app{}'.format(orderLink)
        orderInfosLst.append(orderUrl)
    orderDict = dict(zip(orderNumbersLst, orderInfosLst))
    for orderKey, orderDetailUrl in orderDict.items():
        print(orderKey)
# convert start===========================
        orderDetail = orderDetailUrl
        r = session_requests.get(orderDetail)
        soup = bs(r.text, 'html.parser')
        td_tags = soup.find_all('td')
        rowCount = 0
        colCount = 0
        num = 0
        ws = wb.create_sheet(orderKey)
        ws.append(['編號', '日期', '房號', '房價', '姓名', '身份證字號', '聯絡電話', '出生年月日', '補助金額', '人數'])
        def rowData(rowName):
            cell = '{}{}'.format(rowName,rowCount+2)
            ws[cell].alignment = Alignment(wrapText=True)
            ws[cell].value = '\n'.join(dataLst)
            global colCount
            colCount += 1
        def rowDataBlank():
            global colCount
            colCount += 1
        colNames = {'a': 6, 'b': 8, 'c':8, 'd':8, 'e':12, 'f':15, 'g':15, 'h':12, 'i':10, 'j':6}
        for key, value in colNames.items():
            ws.column_dimensions[key].width = value
# ==============================================================================================
        for td in td_tags:
            dataLst = []
            for data in td.contents:
                if data == '\n':
                    continue
                else:
                    data = str(data)
                    data = data.replace('<div>', '').replace('</div>', '').replace('<p>', '').replace('</p>', '').replace('\n', '')
                    dataLst.append(str(data))
            if colCount == 0:
                cell = 'a{}'.format(rowCount+2)
                ws[cell] = int(dataLst[0])
                colCount += 1
                continue
            if colCount == 1:
                rowData('b')
                continue
            if colCount == 2:
                rowData('c')
                continue
            if colCount == 3:
                rowData('d')
                continue
            if colCount == 4:
                rowData('e')
                continue
            if colCount == 5:
                rowData('f')
                continue
            if colCount == 6:
                rowData('g')
                continue
            if colCount == 7:
                rowData('h')
                continue
            if colCount == 8:
                rowData('i')
                continue
            if colCount == 9:
                cell = 'j{}'.format(rowCount+2)
                printUrl = 'https://springtour.app'+dataLst[0].replace('<a href="', '').replace('">列印</a>', '')
                r = session_requests.get(printUrl)
                soup = bs(r.text, 'html.parser')
                tenant = soup.select('#coupon_data .row p:last-of-type')
                tenant = str(tenant[0]).replace('<p>', '').replace('人</p>', '')
                ws[cell].value = int(tenant)
                colCount = 0
            rowCount += 1
            num += 1
        print(orderKey+' complete!')
# =======================================================================================================

    res = len(wb.sheetnames)
    for sheet in wb:
        if sheet.title == 'Sheet' and len(wb.sheetnames) > 1:
            sheetRemove = wb['Sheet']
            wb.remove(sheetRemove)
    for hotelSheet in wb.sheetnames:
        if hotelSheet == 'Sheet':
            wb.save('{} {} {} 沒有資料.xlsx'.format(hotelCount, hotelName, date))
        else:
            wb.save('{} {} {}.xlsx'.format(hotelCount, hotelName, date))
            # hotelCount += 1
# convert end =========================================
else:
    print(r.status_code)