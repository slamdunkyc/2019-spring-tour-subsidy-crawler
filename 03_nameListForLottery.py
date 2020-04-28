# 完成版，訂單分別存到不同活頁簿
import requests, datetime, time
from bs4 import BeautifulSoup as bs
from openpyxl import Workbook
from openpyxl.styles import Alignment
session_requests = requests.session()
loginUrl = 'https://springtour.app/Ajax/loginadmin'
payload = {
    'account': 'xxxxx',
    'password': 'xxxxx'
}
headers = {
    'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
    'Origin': 'https://springtour.app',
    'Referer': 'https://springtour.app/Main/loginAdmin',
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.142 Safari/537.36',
    'X-Requested-With': 'XMLHttpRequest'
}
date = str(datetime.date.today()).replace('-', '')
r = session_requests.post(loginUrl, data=payload, headers=headers)
if r.status_code == requests.codes.ok:
    print('Login Success!')
    hotelCount = 1
    wb = Workbook()
    ws = wb.active
    ws.append(['飯店名稱', '日期', '房號', '房價', '姓名', '身份證字號', '聯絡電話', '出生年月日', '補助金額', '人數'])
    rowCount = 0
    for i in range(1, 33):
        targetUrl = 'https://springtour.app/Gov/statis/{}'.format(i)
        r = session_requests.get(targetUrl, headers = dict(referer = targetUrl))
        soup = bs(r.text, 'html.parser')
        hotelInfos = soup.find_all('td', {'class': 'cal6'})
        for hotelInfo in hotelInfos:
            hotelLink = hotelInfo.find('a').get('href')
            hotelUrl = 'https://springtour.app{}'.format(hotelLink)
            r = session_requests.get(hotelUrl, headers = dict(referer = targetUrl))
            soup = bs(r.text, 'html.parser')
            hotelNameTag = soup.select('#user_panel p')[0]
            hotelName = hotelNameTag.contents[0]
            print(hotelName)
            orderNumbers = soup.find_all('td', {'class': 'cal1'})
            orderInfos = soup.find_all('td', {'class': 'cal6'})
            orderNumbersLst = []
            orderInfosLst = []
            for orderNumber in orderNumbers:
                orderNumber = orderNumber.text
                orderNumbersLst.append(orderNumber)
            for orderInfo in orderInfos:
                orderLink = orderInfo.find('a').get('href')
                orderUrl = 'https://springtour.app{}'.format(orderLink)
                orderInfosLst.append(orderUrl)
            orderDict = dict(zip(orderNumbersLst, orderInfosLst))
            for orderKey, orderDetailUrl in orderDict.items():
                print(orderKey)
# convert start===========================
                orderDetail = orderDetailUrl
                r = session_requests.get(orderDetail)
                soup = bs(r.text, 'html.parser')
                td_tags = soup.find_all('td')
                colCount = 0
                num = 0
                def rowData(rowName):
                    cell = '{}{}'.format(rowName,rowCount+2)
                    ws[cell].alignment = Alignment(wrapText=True)
                    ws[cell].value = '\n'.join(dataLst)
                    global colCount
                    colCount += 1
                def rowDataBlank():
                    global colCount
                    colCount += 1
                colNames = {'a': 6, 'b': 8, 'c':8, 'd':8, 'e':12, 'f':15, 'g':15, 'h':12, 'i':10, 'j':6}
                for key, value in colNames.items():
                    ws.column_dimensions[key].width = value
                th_tags = soup.find_all('th')
                # print(len(th_tags))
#  ================= find house ======================
                if len(th_tags) == 11:
                    for td in td_tags:
                        dataLst = []
                        for data in td.contents:
                            if data == '\n':
                                continue
                            else:
                                data = str(data)
                                data = data.replace('<div>', '').replace('</div>', '').replace('<p>', '').replace('</p>', '').replace('\n', '')
                                dataLst.append(str(data))
                        if colCount == 0:
                            cell = 'a{}'.format(rowCount+2)
                            ws[cell] = str(hotelName)
                            colCount += 1
                            continue
                        if colCount == 1:
                            rowData('b')
                            continue
                        if colCount == 2:
                            rowData('c')
                            continue
                        if colCount == 3:
                            rowData('d')
                            continue
                        if colCount == 4:
                            rowData('e')
                            continue
                        if colCount == 5:
                            rowData('f')
                            continue
                        if colCount == 6:
                            rowData('g')
                            continue
                        if colCount == 7:
                            rowData('h')
                            continue
                        if colCount == 8:
                            rowData('i')
                            continue
                        if colCount == 9:
                            colCount += 1
                            continue
                        if colCount == 10:
                            cell = 'j{}'.format(rowCount+2)
                            printUrl = 'https://springtour.app'+dataLst[0].replace('<a href="', '').replace('">列印</a>', '')
                            r = session_requests.get(printUrl)
                            soup = bs(r.text, 'html.parser')
                            tenant = soup.select('#coupon_data .row p:last-of-type')
                            tenant = str(tenant[0]).replace('<p>', '').replace('人</p>', '')
                            ws[cell].value = int(tenant)
                            colCount = 0
                        rowCount += 1
                        num += 1
                    print(orderKey+' complete!')
#   =================== find Hotel ===================
                else:
                    for td in td_tags:
                        dataLst = []
                        for data in td.contents:
                            if data == '\n':
                                continue
                            else:
                                data = str(data)
                                data = data.replace('<div>', '').replace('</div>', '').replace('<p>', '').replace('</p>', '').replace('\n', '')
                                dataLst.append(str(data))
                        if colCount == 0:
                            cell = 'a{}'.format(rowCount+2)
                            ws[cell] = str(hotelName)
                            colCount += 1
                            continue
                        if colCount == 1:
                            rowData('b')
                            continue
                        if colCount == 2:
                            rowData('c')
                            continue
                        if colCount == 3:
                            rowData('d')
                            continue
                        if colCount == 4:
                            rowData('e')
                            continue
                        if colCount == 5:
                            rowData('f')
                            continue
                        if colCount == 6:
                            rowData('g')
                            continue
                        if colCount == 7:
                            rowData('h')
                            continue
                        if colCount == 8:
                            rowData('i')
                            continue
                        if colCount == 9:
                            cell = 'j{}'.format(rowCount+2)
                            printUrl = 'https://springtour.app'+dataLst[0].replace('<a href="', '').replace('">列印</a>', '')
                            r = session_requests.get(printUrl)
                            soup = bs(r.text, 'html.parser')
                            tenant = soup.select('#coupon_data .row p:last-of-type')
                            tenant = str(tenant[0]).replace('<p>', '').replace('人</p>', '')
                            ws[cell].value = int(tenant)
                            colCount = 0
                        rowCount += 1
                        num += 1
                    print(orderKey+' complete!')
# =======================================================================================================
            hotelCount += 1
            print('Sleep start')
            time.sleep(3)
    wb.save('住客資料彙整表 {}.xlsx'.format(date))
# convert end =========================================
else:
    print(r.status_code)